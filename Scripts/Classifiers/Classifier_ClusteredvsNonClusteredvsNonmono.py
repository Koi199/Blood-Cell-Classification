"""
classifier_3class.py
────────────────────
Reusable library for 3-class single model blood cell classification.
No cascade — one model classifies directly:
  0 = Unclustered  (MCwRBC + MCwoRBC)
  1 = Clustered
  2 = NonMonocyte  (Unusable + Lymphocyte + RBC alone)

Usage:
    from classifier_3class import train, DEFAULT_CONFIG
    import copy

    config = copy.deepcopy(DEFAULT_CONFIG)
    config["architecture"]    = "resnet50"
    config["checkpoint_path"] = "checkpoints_3class/resnet50.pth"

    train(config, notes="resnet50 3class")
"""

import os
import copy
import torch
import torch.nn as nn
from torch.utils.data import Dataset, DataLoader, WeightedRandomSampler
from torchvision import transforms, models
from PIL import Image
from sklearn.model_selection import train_test_split
from sklearn.metrics import (
    classification_report, confusion_matrix, precision_recall_fscore_support
)
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import load_workbook
from datetime import date

import sys
sys.path.append("C:/repos/Blood-Cell-Classification/Scripts/Logging")
from Logger import (
    setup_mlflow, start_run, log_params, log_epoch,
    log_results, log_artifacts, log_confusion_matrix, end_run
)

# ─────────────────────────────────────────────
# DEFAULT CONFIG
# ─────────────────────────────────────────────
DEFAULT_CONFIG = {
    "data_dir":                 "D:/MMA_LabelledData/Sliced",
    "img_size":                 256,
    "batch_size":               32,
    "num_epochs":               20,
    "early_stopping_patience":  5,
    "lr":                       5e-5,
    "weight_decay":             1e-4,
    "num_workers":              4,
    "checkpoint_path":          "checkpoints_3class/resnet34_3class.pth",
    "device":                   "cuda" if torch.cuda.is_available() else "cpu",
    "classification_threshold": 0.5,
    "architecture":             "resnet34",

    "subclass_targets": {
        "Unusable":   600,
        "Lymphocyte": 150,
        "RBCalone":   400,
        "MCwRBC":     900,
        "MCwoRBC":    900,
        "Clustered":  1800,
    }
}

# ─────────────────────────────────────────────
# LABEL MAPS
# ─────────────────────────────────────────────
FINE_MAP = {
    "Monocyte_with_RBC":    "MCwRBC",
    "Monocyte_without_RBC": "MCwoRBC",
    "Clustered_cell":       "Clustered",
    "Unusable":             "Unusable",
    "Lymphocyte":           "Lymphocyte",
    "RBC alone":            "RBCalone",
}

LABEL_MAP = {
    "MCwRBC":     0,  # Unclustered
    "MCwoRBC":    0,  # Unclustered
    "Clustered":  1,  # Clustered
    "Unusable":   2,  # NonMonocyte
    "Lymphocyte": 2,  # NonMonocyte
    "RBCalone":   2,  # NonMonocyte
}

CLASS_NAMES = ["Unclustered", "Clustered", "NonMonocyte"]

# ─────────────────────────────────────────────
# ViT image size handling
# ─────────────────────────────────────────────
VIT_ARCHITECTURES = {"vit_b16", "vit_b32", "vit_l16"}


def get_img_size(architecture, config_img_size=256):
    if architecture.lower() in VIT_ARCHITECTURES:
        print(f"  ViT detected — overriding img_size to 224")
        return 224
    return config_img_size


def make_transforms(img_size):
    train_tf = transforms.Compose([
        transforms.Resize((img_size, img_size)),
        transforms.RandomHorizontalFlip(),
        transforms.RandomVerticalFlip(),
        transforms.RandomRotation(15),
        transforms.ColorJitter(brightness=0.2, contrast=0.2),
        transforms.ToTensor(),
        transforms.Normalize(mean=[0.485, 0.456, 0.406],
                             std=[0.229, 0.224, 0.225]),
    ])
    val_tf = transforms.Compose([
        transforms.Resize((img_size, img_size)),
        transforms.ToTensor(),
        transforms.Normalize(mean=[0.485, 0.456, 0.406],
                             std=[0.229, 0.224, 0.225]),
    ])
    return train_tf, val_tf

# ─────────────────────────────────────────────
# MODEL BUILDER
# ─────────────────────────────────────────────
def build_model(architecture, num_classes=3, freeze_backbone=False):
    arch = architecture.lower()

    if arch == "resnet18":
        model = models.resnet18(weights=models.ResNet18_Weights.DEFAULT)
    elif arch == "resnet34":
        model = models.resnet34(weights=models.ResNet34_Weights.DEFAULT)
    elif arch == "resnet50":
        model = models.resnet50(weights=models.ResNet50_Weights.DEFAULT)
    elif arch == "resnet101":
        model = models.resnet101(weights=models.ResNet101_Weights.DEFAULT)
    elif arch == "resnet152":
        model = models.resnet152(weights=models.ResNet152_Weights.DEFAULT)
    elif arch == "efficientnet_b0":
        model = models.efficientnet_b0(weights=models.EfficientNet_B0_Weights.DEFAULT)
    elif arch == "efficientnet_b1":
        model = models.efficientnet_b1(weights=models.EfficientNet_B1_Weights.DEFAULT)
    elif arch == "efficientnet_b4":
        model = models.efficientnet_b4(weights=models.EfficientNet_B4_Weights.DEFAULT)
    elif arch == "convnext_tiny":
        model = models.convnext_tiny(weights=models.ConvNeXt_Tiny_Weights.DEFAULT)
    elif arch == "convnext_base":
        model = models.convnext_base(weights=models.ConvNeXt_Base_Weights.DEFAULT)
    elif arch == "vit_b16":
        model = models.vit_b_16(weights=models.ViT_B_16_Weights.DEFAULT)
    else:
        raise ValueError(
            f"Unknown architecture: '{architecture}'. "
            f"Supported: resnet18/34/50/101/152, efficientnet_b0/b1/b4, "
            f"convnext_tiny/base, vit_b16"
        )

    if freeze_backbone:
        for param in model.parameters():
            param.requires_grad = False

    if arch.startswith("resnet"):
        in_features = model.fc.in_features
        model.fc = nn.Sequential(nn.Dropout(p=0.4), nn.Linear(in_features, num_classes))
    elif arch.startswith("efficientnet"):
        in_features = model.classifier[1].in_features
        model.classifier = nn.Sequential(nn.Dropout(p=0.4), nn.Linear(in_features, num_classes))
    elif arch.startswith("convnext"):
        in_features = model.classifier[2].in_features
        model.classifier[2] = nn.Linear(in_features, num_classes)
    elif arch == "vit_b16":
        in_features = model.heads.head.in_features
        model.heads.head = nn.Sequential(nn.Dropout(p=0.4), nn.Linear(in_features, num_classes))

    total     = sum(p.numel() for p in model.parameters())
    trainable = sum(p.numel() for p in model.parameters() if p.requires_grad)
    print(f"  Architecture:     {architecture}")
    print(f"  Total params:     {total:,}")
    print(f"  Trainable params: {trainable:,}")

    return model

# ─────────────────────────────────────────────
# DATASET
# ─────────────────────────────────────────────
class BloodCellDataset(Dataset):
    def __init__(self, samples, transform=None):
        self.samples   = samples
        self.transform = transform

    def __len__(self):
        return len(self.samples)

    def __getitem__(self, idx):
        path, label = self.samples[idx]
        img = Image.open(path).convert("RGB")
        if self.transform:
            img = self.transform(img)
        return img, label

# ─────────────────────────────────────────────
# DATA LOADING
# ─────────────────────────────────────────────
def load_samples(data_dir):
    samples = []
    for subfolder, fine_label in FINE_MAP.items():
        folder_path = os.path.join(data_dir, subfolder)
        if not os.path.isdir(folder_path):
            print(f"Warning: folder not found — {folder_path}")
            continue
        training_label = LABEL_MAP[fine_label]
        for fname in os.listdir(folder_path):
            if fname.lower().endswith(('.png', '.jpg', '.jpeg')):
                samples.append((
                    os.path.join(folder_path, fname),
                    fine_label,
                    training_label
                ))
    return samples


def make_weighted_sampler(train_samples, subclass_targets):
    fine_counts = {}
    for _, fine_label, _ in train_samples:
        fine_counts[fine_label] = fine_counts.get(fine_label, 0) + 1

    print("\nNatural class counts in training set:")
    for label, count in sorted(fine_counts.items()):
        target = subclass_targets.get(label, count)
        print(f"  {label}: {count} → {target} ({target/count:.2f}x)")

    sample_weights = np.zeros(len(train_samples), dtype=np.float32)
    for idx, (_, fine_label, _) in enumerate(train_samples):
        natural = fine_counts[fine_label]
        target  = subclass_targets.get(fine_label, natural)
        sample_weights[idx] = target / natural

    total_samples = sum(subclass_targets.values())
    sampler = WeightedRandomSampler(
        weights=torch.tensor(sample_weights),
        num_samples=total_samples,
        replacement=True
    )

    nonmono     = sum(subclass_targets.get(k, 0) for k in ["Unusable", "Lymphocyte", "RBCalone"])
    unclustered = sum(subclass_targets.get(k, 0) for k in ["MCwRBC", "MCwoRBC"])
    clustered   = subclass_targets.get("Clustered", 0)
    print(f"\nEffective samples per epoch:")
    print(f"  NonMonocyte: {nonmono}  |  Unclustered: {unclustered}  |  "
          f"Clustered: {clustered}  |  Total: {total_samples}")

    return sampler, fine_counts


def make_dataloaders(config, train_tf, val_tf):
    all_samples  = load_samples(config["data_dir"])
    strat_labels = [s[2] for s in all_samples]

    train, temp = train_test_split(
        all_samples, test_size=0.30, stratify=strat_labels, random_state=42
    )
    test, val = train_test_split(
        temp, test_size=0.333,
        stratify=[s[2] for s in temp], random_state=42
    )

    print(f"\nDataset split:")
    print(f"  Train: {len(train)} | Test: {len(test)} | Val: {len(val)}")

    sampler, fine_counts = make_weighted_sampler(train, config["subclass_targets"])

    train_loader = DataLoader(
        BloodCellDataset([(p, l) for p, _, l in train], transform=train_tf),
        batch_size=config["batch_size"], sampler=sampler,
        num_workers=config["num_workers"], pin_memory=True
    )
    val_loader = DataLoader(
        BloodCellDataset([(p, l) for p, _, l in val], transform=val_tf),
        batch_size=config["batch_size"], shuffle=False,
        num_workers=config["num_workers"], pin_memory=True
    )
    test_loader = DataLoader(
        BloodCellDataset([(p, l) for p, _, l in test], transform=val_tf),
        batch_size=config["batch_size"], shuffle=False,
        num_workers=config["num_workers"], pin_memory=True
    )

    return train_loader, val_loader, test_loader, fine_counts

# ─────────────────────────────────────────────
# TRAINING LOOP
# ─────────────────────────────────────────────
def train_one_epoch(model, loader, optimizer, criterion, device):
    model.train()
    total_loss, correct, total = 0, 0, 0
    for imgs, labels in loader:
        imgs, labels = imgs.to(device), labels.to(device)
        optimizer.zero_grad()
        outputs     = model(imgs)
        loss        = criterion(outputs, labels)
        loss.backward()
        optimizer.step()
        total_loss += loss.item() * imgs.size(0)
        correct    += (outputs.argmax(1) == labels).sum().item()
        total      += imgs.size(0)
    return total_loss / total, correct / total

# ─────────────────────────────────────────────
# EVALUATE — 3-class accuracy
# ─────────────────────────────────────────────
@torch.no_grad()
def evaluate(model, loader, criterion, device):
    model.eval()
    total_loss, correct, total = 0, 0, 0
    all_preds, all_labels = [], []

    for imgs, labels in loader:
        imgs, labels = imgs.to(device), labels.to(device)
        outputs      = model(imgs)
        total_loss  += criterion(outputs, labels).item() * imgs.size(0)
        preds        = outputs.argmax(1)
        correct     += (preds == labels).sum().item()
        total       += imgs.size(0)
        all_preds.extend(preds.cpu().numpy())
        all_labels.extend(labels.cpu().numpy())

    return total_loss / total, correct / total, all_preds, all_labels

# ─────────────────────────────────────────────
# PLOTS
# ─────────────────────────────────────────────
def save_plots(train_losses, val_losses, train_accs, val_accs,
               true_labels, preds, output_dir):
    os.makedirs(output_dir, exist_ok=True)

    plt.figure(figsize=(10, 5))
    plt.plot(train_losses, label="Train Loss", marker="o", markersize=3)
    plt.plot(val_losses,   label="Val Loss",   marker="o", markersize=3)
    plt.title("3-Class — Training & Validation Loss")
    plt.xlabel("Epoch")
    plt.ylabel("Loss")
    plt.legend()
    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, "loss_curve_3class.png"), dpi=150)
    plt.close()

    plt.figure(figsize=(10, 5))
    plt.plot(train_accs, label="Train Accuracy", marker="o", markersize=3)
    plt.plot(val_accs,   label="Val Accuracy",   marker="o", markersize=3)
    plt.title("3-Class — Training & Validation Accuracy")
    plt.xlabel("Epoch")
    plt.ylabel("Accuracy")
    plt.ylim(0, 1)
    plt.legend()
    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, "accuracy_curve_3class.png"), dpi=150)
    plt.close()

    cm = confusion_matrix(true_labels, preds, labels=[0, 1, 2])
    plt.figure(figsize=(8, 6))
    sns.heatmap(cm, annot=True, fmt="d", cmap="Blues",
                xticklabels=CLASS_NAMES, yticklabels=CLASS_NAMES)
    plt.title("3-Class — Confusion Matrix (counts)")
    plt.ylabel("True Label")
    plt.xlabel("Predicted Label")
    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, "confusion_matrix_3class.png"), dpi=150)
    plt.close()

    print("  ✓ Saved loss, accuracy, and confusion matrix plots")

# ─────────────────────────────────────────────
# EXCEL LOGGING
# ─────────────────────────────────────────────
def log_experiment_excel(config, results,
                         log_path="C:/repos/Blood-Cell-Classification/experiment_log.xlsx",
                         notes=""):
    wb       = load_workbook(log_path)
    ws       = wb.active
    next_row = ws.max_row + 1
    run_num  = next_row - 2
    targets  = config["subclass_targets"]

    nonmono = sum(targets.get(k, 0) for k in ["Unusable", "Lymphocyte", "RBCalone"])
    mono    = sum(targets.get(k, 0) for k in ["MCwRBC", "MCwoRBC", "Clustered"])

    row = [
        run_num, str(date.today()), notes,
        config.get("architecture", "resnet34"), "ImageNet", "No", 0.4,
        "3-class single model",
        70, 10, 20, nonmono, mono,
        targets.get("MCwRBC", 0), targets.get("MCwoRBC", 0), targets.get("Clustered", 0),
        "Equal balance per class", "AdamW",
        config["lr"], config["weight_decay"], config["batch_size"], config["num_epochs"],
        f"CosineAnnealingLR (T_max={config['num_epochs']})",
        f"{config['img_size']}x{config['img_size']}",
        "HFlip, VFlip, Rotation(15°), ColorJitter(b=0.2,c=0.2), ImageNet norm",
        results["test_acc"],
        results["unclust_prec"],  results["unclust_recall"],  results["unclust_f1"],
        results["clust_prec"],    results["clust_recall"],    results["clust_f1"],
        results["nonmono_prec"],  results["nonmono_recall"],  results["nonmono_f1"],
        results["macro_f1"],      results["weighted_f1"],
        results["val_acc"],
    ]

    for col_idx, value in enumerate(row, start=1):
        ws.cell(row=next_row, column=col_idx, value=value)

    wb.save(log_path)
    print(f"  ✓ Run {run_num} logged to {log_path}")

# ─────────────────────────────────────────────
# MAIN TRAIN FUNCTION
# ─────────────────────────────────────────────
def train(config=None, notes=""):
    """
    Train 3-class single model classifier.

    Args:
        config: dict — override any keys from DEFAULT_CONFIG.
                Pass None to use DEFAULT_CONFIG unchanged.
        notes:  string — description logged to MLflow and Excel.

    Example:
        from classifier_3class import train, DEFAULT_CONFIG
        import copy

        config = copy.deepcopy(DEFAULT_CONFIG)
        config["architecture"]    = "resnet50"
        config["checkpoint_path"] = "checkpoints_3class/resnet50.pth"

        train(config, notes="resnet50 3class")
    """
    cfg = copy.deepcopy(DEFAULT_CONFIG)
    if config is not None:
        for k, v in config.items():
            if k == "subclass_targets" and isinstance(v, dict):
                cfg["subclass_targets"].update(v)
            else:
                cfg[k] = v

    device   = cfg["device"]
    arch     = cfg.get("architecture", "resnet34")
    img_size = get_img_size(arch, cfg["img_size"])
    train_tf, val_tf = make_transforms(img_size)

    print(f"\nTraining on: {device}  |  Architecture: {arch}  |  img_size: {img_size}")
    os.makedirs(os.path.dirname(cfg["checkpoint_path"]), exist_ok=True)
    plot_dir = os.path.dirname(cfg["checkpoint_path"])

    setup_mlflow("3Class_Unclustered_Clustered_NonMono")
    start_run(run_name=f"{arch}_{notes}" if notes else arch, notes=notes)
    log_params(cfg)

    train_loader, val_loader, test_loader, fine_counts = make_dataloaders(cfg, train_tf, val_tf)
    model     = build_model(arch, num_classes=3, freeze_backbone=False).to(device)
    criterion = nn.CrossEntropyLoss()
    optimizer = torch.optim.AdamW(model.parameters(), lr=cfg["lr"], weight_decay=cfg["weight_decay"])
    scheduler = torch.optim.lr_scheduler.CosineAnnealingLR(optimizer, T_max=cfg["num_epochs"])

    train_losses, val_losses = [], []
    train_accs,   val_accs   = [], []
    best_val_acc              = 0.0
    epochs_no_improve         = 0

    print("\n── Training ──")
    for epoch in range(1, cfg["num_epochs"] + 1):
        train_loss, train_acc = train_one_epoch(model, train_loader, optimizer, criterion, device)
        val_loss, val_acc, _, _ = evaluate(model, val_loader, criterion, device)
        scheduler.step()

        train_losses.append(train_loss)
        val_losses.append(val_loss)
        train_accs.append(train_acc)
        val_accs.append(val_acc)

        print(f"Epoch {epoch:02d}/{cfg['num_epochs']} | "
              f"Train Loss: {train_loss:.4f}  Acc: {train_acc:.4f} | "
              f"Val Loss: {val_loss:.4f}  Acc: {val_acc:.4f}")

        if val_acc > best_val_acc:
            best_val_acc      = val_acc
            epochs_no_improve = 0
            torch.save(model.state_dict(), cfg["checkpoint_path"])
            print(f"  ✓ New best model saved (val_acc={val_acc:.4f})")
        else:
            epochs_no_improve += 1
            print(f"  No improvement ({epochs_no_improve}/{cfg['early_stopping_patience']})")
            if epochs_no_improve >= cfg["early_stopping_patience"]:
                print(f"\n  Early stopping at epoch {epoch}")
                break

        log_epoch(epoch, train_loss, train_acc, val_loss, val_acc)

    model.load_state_dict(torch.load(cfg["checkpoint_path"]))

    # ── Val evaluation ──
    print("\n── Validation Set Evaluation ──")
    _, val_acc_final, val_preds, val_true = evaluate(model, val_loader, criterion, device)
    print(f"Val Accuracy: {val_acc_final:.4f}\n")
    print(classification_report(val_true, val_preds, target_names=CLASS_NAMES))
    print(confusion_matrix(val_true, val_preds, labels=[0, 1, 2]))

    # ── Test evaluation ──
    print("\n── Test Set Evaluation ──")
    _, test_acc, preds, true_labels = evaluate(model, test_loader, criterion, device)
    print(f"Test Accuracy: {test_acc:.4f}\n")
    print(classification_report(true_labels, preds, target_names=CLASS_NAMES))
    print(confusion_matrix(true_labels, preds, labels=[0, 1, 2]))

    save_plots(train_losses, val_losses, train_accs, val_accs, true_labels, preds, plot_dir)

    # ── Compile results ──
    prec, rec, f1, _ = precision_recall_fscore_support(true_labels, preds, labels=[0, 1, 2])
    _, _, f1w, _     = precision_recall_fscore_support(true_labels, preds, average="weighted")
    _, _, f1m, _     = precision_recall_fscore_support(true_labels, preds, average="macro")

    results = {
        "test_acc":        round(test_acc, 4),
        "unclust_prec":    round(prec[0],  4),
        "unclust_recall":  round(rec[0],   4),
        "unclust_f1":      round(f1[0],    4),
        "clust_prec":      round(prec[1],  4),
        "clust_recall":    round(rec[1],   4),
        "clust_f1":        round(f1[1],    4),
        "nonmono_prec":    round(prec[2],  4),
        "nonmono_recall":  round(rec[2],   4),
        "nonmono_f1":      round(f1[2],    4),
        "macro_f1":        round(f1m,      4),
        "weighted_f1":     round(f1w,      4),
        "val_acc":         round(val_acc_final, 4),
    }

    log_experiment_excel(cfg, results, notes=notes)
    log_results(results)
    log_confusion_matrix(true_labels, preds, CLASS_NAMES,
                         os.path.join(plot_dir, "confusion_matrix_3class.png"))
    log_artifacts([
        os.path.join(plot_dir, "loss_curve_3class.png"),
        os.path.join(plot_dir, "accuracy_curve_3class.png"),
        cfg["checkpoint_path"],
    ])
    end_run()

    return results


# ─────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────
if __name__ == "__main__":
    train(notes="default config run")