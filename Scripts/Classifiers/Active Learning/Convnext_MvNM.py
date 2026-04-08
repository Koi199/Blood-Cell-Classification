# ================================================================
# classifier_stage1_convnext.py
# ConvNeXt classifier with support for custom dataset splits
# ================================================================

import os
import copy
import sys
import json
from datetime import date

import numpy as np
import torch
import torch.nn as nn
from PIL import Image
from sklearn.metrics import (
    classification_report,
    confusion_matrix,
    precision_recall_fscore_support,
)
from torch.utils.data import Dataset, DataLoader, WeightedRandomSampler
from torchvision import transforms, models
from sklearn.model_selection import train_test_split
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import load_workbook
from datetime import date

# MLflow logger
sys.path.append("C:/repos/Blood-Cell-Classification/Scripts/Logging")
from Logger import (
    setup_mlflow,
    start_run,
    log_params,
    log_epoch,
    log_results,
    log_artifacts,
    log_confusion_matrix,
    end_run,
)

# ================================================================
# DEFAULT CONFIG
# ================================================================
_mono_total = 2450

DEFAULT_CONFIG = {
    "data_dir": "D:/MMA_LabelledData/Sliced",
    "split_json": None,  # path to dataset_splits.json

    "img_size": 256,
    "batch_size": 32,
    "num_epochs": 30,
    "early_stopping_patience": 7,
    "lr": 1e-5,
    "weight_decay": 1e-4,
    "num_workers": 4,
    "checkpoint_path": "checkpoints_stage1/convnext_tiny.pth",
    "device": "cuda" if torch.cuda.is_available() else "cpu",
    "classification_threshold": 0.45,
    "architecture": "convnext_tiny",

    "subclass_targets": {
        "Unusable": 2050,
        "RBCalone": 400,
        "MCwRBC": _mono_total // 3,
        "MCwoRBC": _mono_total // 3,
        "Clustered": _mono_total - 2 * (_mono_total // 3),
    }
}

# ================================================================
# LABEL MAPS
# ================================================================
FINE_MAP = {
    "Monocyte_with_RBC": "MCwRBC",
    "Monocyte_without_RBC": "MCwoRBC",
    "Clustered_cell": "Clustered",
    "Unusable": "Unusable",
    "RBC alone": "RBCalone",
}

LABEL_MAP = {
    "Unusable": 0,
    "RBCalone": 1,
    "MCwRBC": 2,
    "MCwoRBC": 2,
    "Clustered": 2,
}

COLLAPSE_MAP = {0: 0, 1: 0, 2: 1}

# ================================================================
# MODEL BUILDER
# ================================================================
def build_model(architecture="convnext_tiny", num_classes=3, freeze_backbone=False):
    arch = architecture.lower()

    if arch == "convnext_tiny":
        model = models.convnext_tiny(weights=models.ConvNeXt_Tiny_Weights.DEFAULT)
    elif arch == "convnext_base":
        model = models.convnext_base(weights=models.ConvNeXt_Base_Weights.DEFAULT)
    else:
        raise ValueError("Only convnext_tiny or convnext_base supported")

    if freeze_backbone:
        for p in model.parameters():
            p.requires_grad = False

    in_features = model.classifier[2].in_features
    model.classifier[2] = nn.Linear(in_features, num_classes)

    return model

# ================================================================
# TRANSFORMS
# ================================================================
base_train_transforms = transforms.Compose([
    transforms.Resize((256, 256)),
    transforms.RandomHorizontalFlip(),
    transforms.RandomVerticalFlip(),
    transforms.RandomRotation(15),
    transforms.ColorJitter(brightness=0.2, contrast=0.2),
    transforms.ToTensor(),
    transforms.Normalize(mean=[0.485, 0.456, 0.406],
                         std=[0.229, 0.224, 0.225]),
])

oversample_transforms = transforms.Compose([
    transforms.Resize((256, 256)),
    transforms.RandomHorizontalFlip(),
    transforms.RandomVerticalFlip(),
    transforms.RandomRotation(25),
    transforms.RandomApply([
        transforms.ColorJitter(brightness=0.4, contrast=0.4, saturation=0.2)
    ], p=0.5),
    transforms.RandomApply([
        transforms.GaussianBlur(kernel_size=3)
    ], p=0.3),
    transforms.RandomApply([
        transforms.RandomAffine(degrees=0, translate=(0.1, 0.1))
    ], p=0.3),
    transforms.ToTensor(),
    transforms.Normalize(mean=[0.485, 0.456, 0.406],
                         std=[0.229, 0.224, 0.225]),
])

val_transforms = transforms.Compose([
    transforms.Resize((256, 256)),
    transforms.ToTensor(),
    transforms.Normalize(mean=[0.485, 0.456, 0.406],
                         std=[0.229, 0.224, 0.225]),
])

# ================================================================
# DATASET
# ================================================================
class BloodCellDataset(Dataset):
    def __init__(self, samples, base_transform, oversample_transform, oversampled_classes):
        self.samples = samples
        self.base_transform = base_transform
        self.oversample_transform = oversample_transform
        self.oversampled_classes = oversampled_classes

    def __len__(self):
        return len(self.samples)

    def __getitem__(self, idx):
        path, training_label, fine_label = self.samples[idx]
        img = Image.open(path).convert("RGB")

        if fine_label in self.oversampled_classes:
            img = self.oversample_transform(img)
        else:
            img = self.base_transform(img)

        return img, training_label

# ================================================================
# CUSTOM SPLIT SUPPORT
# ================================================================
def resolve_label_for_filename(fname, data_dir):
    """Find which class folder contains fname."""
    for folder in os.listdir(data_dir):
        folder_path = os.path.join(data_dir, folder)
        if not os.path.isdir(folder_path):
            continue

        candidate = os.path.join(folder_path, fname)
        if os.path.isfile(candidate):
            fine_label = FINE_MAP.get(folder, folder)
            training_label = LABEL_MAP[fine_label]
            return candidate, training_label, fine_label

    raise FileNotFoundError(f"{fname} not found in {data_dir}")

def build_samples_from_filenames(filenames, data_dir):
    return [resolve_label_for_filename(fname, data_dir) for fname in filenames]

# ================================================================
# SAMPLER
# ================================================================
def make_weighted_sampler(train_samples, subclass_targets):
    fine_counts = {}
    for _, _, fine_label in train_samples:
        fine_counts[fine_label] = fine_counts.get(fine_label, 0) + 1

    sample_weights = np.zeros(len(train_samples), dtype=np.float32)
    for idx, (_, _, fine_label) in enumerate(train_samples):
        natural = fine_counts[fine_label]
        target = subclass_targets.get(fine_label, natural)
        sample_weights[idx] = target / natural

    total_samples = sum(subclass_targets.values())
    sampler = WeightedRandomSampler(
        weights=torch.tensor(sample_weights),
        num_samples=total_samples,
        replacement=True
    )

    return sampler, fine_counts

# ================================================================
# DATALOADERS
# ================================================================
def make_dataloaders(config):
    if config.get("custom_train_samples") is not None:
        train = config["custom_train_samples"]
        val = config["custom_val_samples"]
        test = config["custom_test_samples"]
    else:
        raise RuntimeError("No custom split provided. Set cfg['split_json'].")

    sampler, fine_counts = make_weighted_sampler(train, config["subclass_targets"])

    oversampled_classes = {
        k for k, v in config["subclass_targets"].items()
        if v > fine_counts.get(k, 0)
    }

    train_loader = DataLoader(
        BloodCellDataset(train, base_train_transforms, oversample_transforms, oversampled_classes),
        batch_size=config["batch_size"],
        sampler=sampler,
        num_workers=config["num_workers"],
        pin_memory=True
    )

    val_loader = DataLoader(
        BloodCellDataset(val, val_transforms, val_transforms, set()),
        batch_size=config["batch_size"],
        shuffle=False,
        num_workers=config["num_workers"],
        pin_memory=True
    )

    test_loader = DataLoader(
        BloodCellDataset(test, val_transforms, val_transforms, set()),
        batch_size=config["batch_size"],
        shuffle=False,
        num_workers=config["num_workers"],
        pin_memory=True
    )

    return train_loader, val_loader, test_loader, fine_counts

# ─────────────────────────────────────────────
# MULTICLASS DIAGNOSTIC
# ─────────────────────────────────────────────
def evaluate_multiclass(model, loader, device, output_dir):
    model.eval()
    all_preds, all_labels = [], []
    with torch.no_grad():
        for imgs, labels in loader:
            imgs = imgs.to(device)
            all_preds.extend(model(imgs).argmax(1).cpu().numpy())
            all_labels.extend(labels.cpu().numpy())

    cm = confusion_matrix(all_labels, all_preds, labels=[0, 1, 2])
    print("\n── 3-Class Confusion Matrix (raw counts) ──")
    print(cm)

    plt.figure(figsize=(8, 6))
    sns.heatmap(cm, annot=True, fmt="d", cmap="Blues",
                xticklabels=["Unusable/Lymphocyte", "RBCalone", "Monocyte"],
                yticklabels=["Unusable/Lymphocyte", "RBCalone", "Monocyte"])
    plt.title("3-Class Confusion Matrix (counts)")
    plt.ylabel("True Label")
    plt.xlabel("Predicted Label")
    plt.tight_layout()
    path = os.path.join(output_dir, "confusion_matrix_3class.png")
    plt.savefig(path, dpi=150)
    plt.close()
    print(f"  ✓ Saved {path}")

# ─────────────────────────────────────────────
# THRESHOLD SEARCH
# ─────────────────────────────────────────────
def threshold_search(model, test_loader, device):
    all_probs, all_labels = [], []
    model.eval()
    with torch.no_grad():
        for imgs, labels in test_loader:
            imgs = imgs.to(device)
            all_probs.extend(torch.softmax(model(imgs), dim=1)[:, 2].cpu().numpy())
            all_labels.extend([COLLAPSE_MAP[l.item()] for l in labels])

    print(f"\n── Threshold Search ──")
    print(f"{'Threshold':>10} {'Mono Recall':>12} {'Mono Prec':>10} "
          f"{'NonMono Recall':>15} {'NonMono Prec':>13} {'Accuracy':>10}")
    print("-" * 75)

    for t in np.arange(0.20, 0.81, 0.05):
        preds        = (np.array(all_probs) >= t).astype(int)
        prec, rec, _, _ = precision_recall_fscore_support(
            all_labels, preds, labels=[0, 1], zero_division=0
        )
        acc = np.mean(preds == np.array(all_labels))
        print(f"{t:>10.2f} {rec[1]:>12.4f} {prec[1]:>10.4f} "
              f"{rec[0]:>15.4f} {prec[0]:>13.4f} {acc:>10.4f}")

    print(f"\n  Set classification_threshold in config manually.")

# ================================================================
# TRAINING LOOP
# ================================================================
def _train_one_epoch_clean(model, loader, optimizer, criterion, device):
    model.train()
    total_loss, correct, total = 0, 0, 0

    for imgs, labels in loader:
        imgs, labels = imgs.to(device), labels.to(device)
        optimizer.zero_grad()
        outputs = model(imgs)
        loss = criterion(outputs, labels)
        loss.backward()
        optimizer.step()

        total_loss += loss.item() * imgs.size(0)
        correct += (outputs.argmax(1) == labels).sum().item()
        total += imgs.size(0)

    return total_loss / total, correct / total

# ================================================================
# EVALUATION
# ================================================================
@torch.no_grad()
def evaluate(model, loader, criterion, device, threshold=0.5):
    model.eval()
    total_loss, total = 0, 0
    all_preds, all_labels = [], []

    for imgs, labels in loader:
        imgs, labels = imgs.to(device), labels.to(device)
        outputs = model(imgs)

        total_loss += criterion(outputs, labels).item() * imgs.size(0)
        total += imgs.size(0)

        mono_prob = torch.softmax(outputs, dim=1)[:, 2]
        binary_preds = (mono_prob >= threshold).long()
        binary_true = torch.tensor(
            [COLLAPSE_MAP[l.item()] for l in labels], device=device
        )

        all_preds.extend(binary_preds.cpu().numpy())
        all_labels.extend(binary_true.cpu().numpy())

    acc = sum(p == l for p, l in zip(all_preds, all_labels)) / total
    return total_loss / total, acc, all_preds, all_labels

# ─────────────────────────────────────────────
# PLOTS
# ─────────────────────────────────────────────
def save_plots(train_losses, val_losses, train_accs, val_accs,
               true_labels, preds, output_dir):
    os.makedirs(output_dir, exist_ok=True)

    plt.figure(figsize=(10, 5))
    plt.plot(train_losses, label="Train Loss", marker="o", markersize=3)
    plt.plot(val_losses,   label="Val Loss",   marker="o", markersize=3)
    plt.title("Training & Validation Loss")
    plt.xlabel("Epoch")
    plt.ylabel("Loss")
    plt.legend()
    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, "loss_curve.png"), dpi=150)
    plt.close()

    plt.figure(figsize=(10, 5))
    plt.plot(train_accs, label="Train Accuracy (3-class)", marker="o", markersize=3)
    plt.plot(val_accs,   label="Val Accuracy (binary)",    marker="o", markersize=3)
    plt.title("Training & Validation Accuracy")
    plt.xlabel("Epoch")
    plt.ylabel("Accuracy")
    plt.ylim(0, 1)
    plt.legend()
    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, "accuracy_curve.png"), dpi=150)
    plt.close()

    print("  ✓ Saved loss_curve.png and accuracy_curve.png")

# ─────────────────────────────────────────────
# EXCEL LOGGING
# ─────────────────────────────────────────────
def log_experiment_excel(config, results,
                         log_path="C:/repos/Blood-Cell-Classification/ExperimentLog/experiment_log.xlsx",
                         notes=""):
    wb      = load_workbook(log_path)
    ws      = wb.active
    next_row = ws.max_row + 1
    run_num  = next_row - 2
    targets  = config["subclass_targets"]

    nonmono = sum(targets.get(k, 0) for k in ["Unusable", "Lymphocyte", "RBCalone"])
    mono    = sum(targets.get(k, 0) for k in ["MCwRBC", "MCwoRBC", "Clustered"])

    row = [
        run_num, str(date.today()), notes,
        config.get("architecture", "resnet34"), "ImageNet", "No", 0.4, 12494,
        70, 10, 20, nonmono, mono,
        targets.get("MCwRBC", 0), targets.get("MCwoRBC", 0), targets.get("Clustered", 0),
        "Per-subclass inverse frequency", "AdamW",
        config["lr"], config["weight_decay"], config["batch_size"], config["num_epochs"],
        f"CosineAnnealingLR (T_max={config['num_epochs']})",
        f"{config['img_size']}x{config['img_size']}",
        "HFlip, VFlip, Rotation(15°), ColorJitter(b=0.2,c=0.2), ImageNet norm",
        results["test_acc"],
        results["non_mono_prec"], results["non_mono_recall"], results["non_mono_f1"],
        results["mono_prec"],     results["mono_recall"],     results["mono_f1"],
        results["macro_f1"],      results["weighted_f1"],
        results["tp"], results["tn"], results["fp"], results["fn"],
        results["val_acc"],
    ]

    for col_idx, value in enumerate(row, start=1):
        ws.cell(row=next_row, column=col_idx, value=value)

    wb.save(log_path)
    print(f"  ✓ Run {run_num} logged to {log_path}")

# ================================================================
# MAIN TRAIN FUNCTION
# ================================================================
def train(config=None, notes=""):
    cfg = copy.deepcopy(DEFAULT_CONFIG)
    if config is not None:
        cfg.update(config)

    device = cfg["device"]
    arch = cfg["architecture"]

    os.makedirs(os.path.dirname(cfg["checkpoint_path"]), exist_ok=True)
    plot_dir = os.path.dirname(cfg["checkpoint_path"])

    # # Load custom split
    # if cfg["split_json"] is None:
    #     raise RuntimeError("You must set cfg['split_json'] to dataset_splits.json")

    # with open(cfg["split_json"], "r") as f:
    #     split = json.load(f)

    # cfg["custom_train_samples"] = build_samples_from_filenames(split["train"], cfg["data_dir"])
    # cfg["custom_val_samples"] = build_samples_from_filenames(split["val"], cfg["data_dir"])
    # cfg["custom_test_samples"] = build_samples_from_filenames(split["test"], cfg["data_dir"])

    # If AL driver provides custom samples, use them
    if cfg.get("custom_train_samples") is None:
        raise RuntimeError("Active Learning requires custom_train_samples to be provided.")

    # Logging
    setup_mlflow("Stage1_MCvsNonMC")
    start_run(run_name=f"{arch}_{notes}", notes=notes)
    log_params(cfg)

    # Dataloaders
    train_loader, val_loader, test_loader, fine_counts = make_dataloaders(cfg)

    # Model
    model = build_model(arch, num_classes=3).to(device)
    criterion = nn.CrossEntropyLoss()
    optimizer = torch.optim.AdamW(model.parameters(), lr=cfg["lr"], weight_decay=cfg["weight_decay"])
    scheduler = torch.optim.lr_scheduler.CosineAnnealingLR(optimizer, T_max=cfg["num_epochs"])

    # Training loop
    train_losses, val_losses = [], []
    train_accs, val_accs = [],[]
    best_val_acc = 0
    epochs_no_improve = 0

    for epoch in range(1, cfg["num_epochs"] + 1):
        train_loss, train_acc = _train_one_epoch_clean(model, train_loader, optimizer, criterion, device)
        val_loss, val_acc, _, _ = evaluate(model, val_loader, criterion, device, cfg["classification_threshold"])
        scheduler.step()

        train_losses.append(train_loss)
        val_losses.append(val_loss)
        train_accs.append(train_acc)
        val_accs.append(val_acc)

        print(f"Epoch {epoch:02d}/{cfg['num_epochs']} | "
              f"Train Loss: {train_loss:.4f}  Acc: {train_acc:.4f} | "
              f"Val Loss: {val_loss:.4f}  Acc: {val_acc:.4f}")

        if val_acc > best_val_acc:
            best_val_acc      = val_acc
            epochs_no_improve = 0
            torch.save(model.state_dict(), cfg["checkpoint_path"])
            print(f"  ✓ New best model saved (val_acc={val_acc:.4f})")
        else:
            epochs_no_improve += 1
            print(f"  No improvement ({epochs_no_improve}/{cfg['early_stopping_patience']})")
            if epochs_no_improve >= cfg["early_stopping_patience"]:
                print(f"\n  Early stopping at epoch {epoch}")
                break

        log_epoch(epoch, train_loss, train_acc, val_loss, val_acc)

    # ── Load best checkpoint ──
    model.load_state_dict(torch.load(cfg["checkpoint_path"]))

    # ── Diagnostics ──
    evaluate_multiclass(model, test_loader, device, plot_dir)
    threshold_search(model, test_loader, device)

    # ── Val evaluation ──
    print("\n── Validation Set Evaluation ──")
    _, val_acc_final, val_preds, val_true = evaluate(
        model, val_loader, criterion, device, cfg["classification_threshold"]
    )
    print(f"Val Accuracy: {val_acc_final:.4f}\n")
    print(classification_report(val_true, val_preds, target_names=["Non-Monocyte", "Monocyte"]))
    print(confusion_matrix(val_true, val_preds))

    # ── Test evaluation ──
    print("\n── Test Set Evaluation ──")
    _, test_acc, preds, true_labels = evaluate(
        model, test_loader, criterion, device, cfg["classification_threshold"]
    )
    print(f"Test Accuracy: {test_acc:.4f}\n")
    print(classification_report(true_labels, preds, target_names=["Non-Monocyte", "Monocyte"]))
    print(confusion_matrix(true_labels, preds))

    # ── Save plots ──
    save_plots(train_losses, val_losses, train_accs, val_accs, true_labels, preds, plot_dir)

    # ── Compile results ──
    prec, rec, f1, _ = precision_recall_fscore_support(true_labels, preds, labels=[0, 1])
    _, _, f1w, _     = precision_recall_fscore_support(true_labels, preds, average="weighted")
    _, _, f1m, _     = precision_recall_fscore_support(true_labels, preds, average="macro")
    cm               = confusion_matrix(true_labels, preds)

    results = {
        "test_acc":        round(test_acc,   4),
        "non_mono_prec":   round(prec[0],    4),
        "non_mono_recall": round(rec[0],     4),
        "non_mono_f1":     round(f1[0],      4),
        "mono_prec":       round(prec[1],    4),
        "mono_recall":     round(rec[1],     4),
        "mono_f1":         round(f1[1],      4),
        "macro_f1":        round(f1m,        4),
        "weighted_f1":     round(f1w,        4),
        "tp": int(cm[1,1]), "tn": int(cm[0,0]),
        "fp": int(cm[0,1]), "fn": int(cm[1,0]),
        "val_acc": round(val_acc_final, 4),
    }

    # ── Log everything ──
    log_experiment_excel(cfg, results, notes=notes)
    log_results(results)
    log_confusion_matrix(true_labels, preds, ["Non-Monocyte", "Monocyte"],
                         os.path.join(plot_dir, "confusion_matrix.png"))
    log_artifacts([
        os.path.join(plot_dir, "loss_curve.png"),
        os.path.join(plot_dir, "accuracy_curve.png"),
        os.path.join(plot_dir, "confusion_matrix_3class.png"),
        cfg["checkpoint_path"],
    ])
    end_run()

    return results

# ================================================================
# MAIN
# ================================================================
if __name__ == "__main__":
    cfg = {
        "split_json": "D:/MMA_LabelledData/splits/dataset_splits.json",
        "checkpoint_path": "checkpoints_stage1/convnext_tiny.pth"
    }
    train(cfg, notes="Stage1_fixed_split_Ignoring Lymphocyte class")
