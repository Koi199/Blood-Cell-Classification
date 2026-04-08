"""
classifier_stage1.py
────────────────────
Reusable library for Stage 1 blood cell classification.
Supports pluggable model architectures and configurable sampling targets.

Usage:
    from classifier_stage1 import train, DEFAULT_CONFIG, build_model
    import copy

    config = copy.deepcopy(DEFAULT_CONFIG)
    config["architecture"]    = "resnet50"
    config["checkpoint_path"] = "checkpoints_stage1/resnet50.pth"
    config["subclass_targets"]["MCwRBC"] = 900

    train(config, notes="resnet50 higher MCwRBC target")
"""

import os
import copy
import torch
import torch.nn as nn
from torch.utils.data import Dataset, DataLoader, WeightedRandomSampler
from torchvision import transforms, models
from PIL import Image
from sklearn.model_selection import train_test_split
from sklearn.metrics import (
    classification_report, confusion_matrix, precision_recall_fscore_support
)
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import load_workbook
from datetime import date

import sys
sys.path.append("C:/repos/Blood-Cell-Classification/Scripts/Logging")
from Logger import (
    setup_mlflow, start_run, log_params, log_epoch,
    log_results, log_artifacts, log_confusion_matrix, end_run
)

# ─────────────────────────────────────────────
# DEFAULT CONFIG
# Override any key when calling train()
# ─────────────────────────────────────────────
_mono_total = 2421

DEFAULT_CONFIG = {
    "data_dir":                 "D:/MMA_LabelledData/Sliced",
    "img_size":                 256,
    "batch_size":               32,
    "num_epochs":               20,
    "early_stopping_patience":  5,
    "lr":                       5e-5,
    "weight_decay":             1e-4,
    "num_workers":              4,
    "checkpoint_path":          "checkpoints_stage1/resnet34_binary.pth",
    "device":                   "cuda" if torch.cuda.is_available() else "cpu",
    "classification_threshold": 0.45,
    "architecture":             "resnet34",   # options: resnet18/34/50/101/152, efficientnet_b0/b1/b4, convnext_tiny/base, vit_b16

    "subclass_targets": {
        "Unusable":   1800,
        "RBCalone":   400,
        "MCwRBC":     _mono_total // 3,
        "MCwoRBC":    _mono_total // 3,
        "Clustered":  _mono_total - 2 * (_mono_total // 3),
    }
}

# ─────────────────────────────────────────────
# LABEL MAPS
# ─────────────────────────────────────────────
FINE_MAP = {
    "Monocyte_with_RBC":    "MCwRBC",
    "Monocyte_without_RBC": "MCwoRBC",
    "Clustered_cell":       "Clustered",
    "Unusable":             "Unusable",
    "RBC alone":            "RBCalone",
}

LABEL_MAP = {
    "Unusable":   0,
    "RBCalone":   1,
    "MCwRBC":     2,
    "MCwoRBC":    2,
    "Clustered":  2,
}

COLLAPSE_MAP = {
    0: 0,
    1: 0,
    2: 1,
}

# ─────────────────────────────────────────────
# MODEL BUILDER
# Supports multiple architectures via config["architecture"]
# ─────────────────────────────────────────────
def build_model(architecture, num_classes=3, freeze_backbone=False):
    """
    Build a pretrained model with a custom classification head.

    Supported architectures:
        resnet18, resnet34, resnet50, resnet101, resnet152
        efficientnet_b0, efficientnet_b1, efficientnet_b4
        convnext_tiny, convnext_base
        vit_b16

    Args:
        architecture:    string name from supported list above
        num_classes:     number of output classes
        freeze_backbone: freeze all layers except the head

    Returns:
        model: PyTorch model ready for training
    """
    arch = architecture.lower()

    # ── ResNet family ──
    if arch == "resnet18":
        model = models.resnet18(weights=models.ResNet18_Weights.DEFAULT)
    elif arch == "resnet34":
        model = models.resnet34(weights=models.ResNet34_Weights.DEFAULT)
    elif arch == "resnet50":
        model = models.resnet50(weights=models.ResNet50_Weights.DEFAULT)
    elif arch == "resnet101":
        model = models.resnet101(weights=models.ResNet101_Weights.DEFAULT)
    elif arch == "resnet152":
        model = models.resnet152(weights=models.ResNet152_Weights.DEFAULT)

    # ── EfficientNet family ──
    elif arch == "efficientnet_b0":
        model = models.efficientnet_b0(weights=models.EfficientNet_B0_Weights.DEFAULT)
    elif arch == "efficientnet_b1":
        model = models.efficientnet_b1(weights=models.EfficientNet_B1_Weights.DEFAULT)
    elif arch == "efficientnet_b4":
        model = models.efficientnet_b4(weights=models.EfficientNet_B4_Weights.DEFAULT)

    # ── ConvNeXt family ──
    elif arch == "convnext_tiny":
        model = models.convnext_tiny(weights=models.ConvNeXt_Tiny_Weights.DEFAULT)
    elif arch == "convnext_base":
        model = models.convnext_base(weights=models.ConvNeXt_Base_Weights.DEFAULT)

    # ── Vision Transformer ──
    elif arch == "vit_b16":
        model = models.vit_b_16(weights=models.ViT_B_16_Weights.DEFAULT)

    else:
        raise ValueError(
            f"Unknown architecture: '{architecture}'. "
            f"Supported: resnet18/34/50/101/152, efficientnet_b0/b1/b4, "
            f"convnext_tiny/base, vit_b16"
        )

    if freeze_backbone:
        for param in model.parameters():
            param.requires_grad = False

    # ── Replace classification head ──
    if arch.startswith("resnet"):
        in_features = model.fc.in_features
        model.fc = nn.Sequential(
            nn.Dropout(p=0.4),
            nn.Linear(in_features, num_classes)
        )

    elif arch.startswith("efficientnet"):
        in_features = model.classifier[1].in_features
        model.classifier = nn.Sequential(
            nn.Dropout(p=0.4),
            nn.Linear(in_features, num_classes)
        )

    elif arch.startswith("convnext"):
        in_features = model.classifier[2].in_features
        model.classifier[2] = nn.Linear(in_features, num_classes)

    elif arch == "vit_b16":
        in_features = model.heads.head.in_features
        model.heads.head = nn.Sequential(
            nn.Dropout(p=0.4),
            nn.Linear(in_features, num_classes)
        )

    total_params     = sum(p.numel() for p in model.parameters())
    trainable_params = sum(p.numel() for p in model.parameters() if p.requires_grad)
    print(f"  Architecture:     {architecture}")
    print(f"  Total params:     {total_params:,}")
    print(f"  Trainable params: {trainable_params:,}")

    return model

# ─────────────────────────────────────────────
# DATASET
# ─────────────────────────────────────────────
class BloodCellDataset(Dataset):
    def __init__(self, samples, transform=None):
        self.samples   = samples
        self.transform = transform

    def __len__(self):
        return len(self.samples)

    def __getitem__(self, idx):
        path, label = self.samples[idx]
        img = Image.open(path).convert("RGB")
        if self.transform:
            img = self.transform(img)
        return img, label

# ─────────────────────────────────────────────
# TRANSFORMS
# ─────────────────────────────────────────────
train_transforms = transforms.Compose([
    transforms.Resize((256, 256)),
    transforms.RandomHorizontalFlip(),
    transforms.RandomVerticalFlip(),
    transforms.RandomRotation(15),
    transforms.ColorJitter(brightness=0.2, contrast=0.2),
    transforms.ToTensor(),
    transforms.Normalize(mean=[0.485, 0.456, 0.406],
                         std=[0.229, 0.224, 0.225]),
])

val_transforms = transforms.Compose([
    transforms.Resize((256, 256)),
    transforms.ToTensor(),
    transforms.Normalize(mean=[0.485, 0.456, 0.406],
                         std=[0.229, 0.224, 0.225]),
])

# ─────────────────────────────────────────────
# DATA LOADING
# ─────────────────────────────────────────────
def load_samples(data_dir):
    samples = []
    for subfolder, fine_label in FINE_MAP.items():
        folder_path = os.path.join(data_dir, subfolder)
        if not os.path.isdir(folder_path):
            print(f"Warning: folder not found — {folder_path}")
            continue
        training_label = LABEL_MAP[fine_label]
        for fname in os.listdir(folder_path):
            if fname.lower().endswith(('.png', '.jpg', '.jpeg')):
                samples.append((
                    os.path.join(folder_path, fname),
                    fine_label,
                    training_label
                ))
    return samples


def make_weighted_sampler(train_samples, subclass_targets):
    fine_counts = {}
    for _, fine_label, _ in train_samples:
        fine_counts[fine_label] = fine_counts.get(fine_label, 0) + 1

    print("\nNatural class counts in training set:")
    for label, count in sorted(fine_counts.items()):
        target = subclass_targets.get(label, count)
        binary = "Monocyte" if LABEL_MAP.get(label, 0) == 2 else "Non-Monocyte"
        print(f"  {label} ({binary}): {count} → {target} ({target/count:.2f}x)")

    sample_weights = np.zeros(len(train_samples), dtype=np.float32)
    for idx, (_, fine_label, _) in enumerate(train_samples):
        natural = fine_counts[fine_label]
        target  = subclass_targets.get(fine_label, natural)
        sample_weights[idx] = target / natural

    total_samples = sum(subclass_targets.values())
    sampler = WeightedRandomSampler(
        weights=torch.tensor(sample_weights),
        num_samples=total_samples,
        replacement=True
    )

    nonmono = sum(subclass_targets.get(k, 0) for k in ["Unusable", "RBCalone"])
    mono    = sum(subclass_targets.get(k, 0) for k in ["MCwRBC", "MCwoRBC", "Clustered"])
    print(f"\nEffective samples per epoch:")
    print(f"  NonMonocyte: {nonmono}  |  Monocyte: {mono}  |  Total: {total_samples}")

    return sampler, fine_counts


def make_dataloaders(config):
    all_samples   = load_samples(config["data_dir"])
    binary_labels = [COLLAPSE_MAP[s[2]] for s in all_samples]

    train, temp = train_test_split(
        all_samples, test_size=0.30, stratify=binary_labels, random_state=42
    )
    test, val = train_test_split(
        temp, test_size=0.333,
        stratify=[COLLAPSE_MAP[s[2]] for s in temp], random_state=42
    )

    print(f"\nDataset split:")
    print(f"  Train: {len(train)} | Test: {len(test)} | Val: {len(val)}")

    sampler, fine_counts = make_weighted_sampler(train, config["subclass_targets"])

    train_loader = DataLoader(
        BloodCellDataset([(p, l) for p, _, l in train], transform=train_transforms),
        batch_size=config["batch_size"], sampler=sampler,
        num_workers=config["num_workers"], pin_memory=True
    )
    val_loader = DataLoader(
        BloodCellDataset([(p, l) for p, _, l in val], transform=val_transforms),
        batch_size=config["batch_size"], shuffle=False,
        num_workers=config["num_workers"], pin_memory=True
    )
    test_loader = DataLoader(
        BloodCellDataset([(p, l) for p, _, l in test], transform=val_transforms),
        batch_size=config["batch_size"], shuffle=False,
        num_workers=config["num_workers"], pin_memory=True
    )

    return train_loader, val_loader, test_loader, fine_counts

# ─────────────────────────────────────────────
# TRAINING LOOP
# ─────────────────────────────────────────────
def _train_one_epoch_clean(model, loader, optimizer, criterion, device):
    model.train()
    total_loss, correct, total = 0, 0, 0
    for imgs, labels in loader:
        imgs, labels = imgs.to(device), labels.to(device)
        optimizer.zero_grad()
        outputs = model(imgs)
        loss    = criterion(outputs, labels)
        loss.backward()
        optimizer.step()
        total_loss += loss.item() * imgs.size(0)
        correct    += (outputs.argmax(1) == labels).sum().item()
        total      += imgs.size(0)
    return total_loss / total, correct / total

# ─────────────────────────────────────────────
# EVALUATE
# ─────────────────────────────────────────────
@torch.no_grad()
def evaluate(model, loader, criterion, device, threshold=0.5):
    model.eval()
    total_loss, total = 0, 0
    all_preds, all_labels = [], []

    for imgs, labels in loader:
        imgs, labels = imgs.to(device), labels.to(device)
        outputs      = model(imgs)
        total_loss  += criterion(outputs, labels).item() * imgs.size(0)
        total       += imgs.size(0)

        mono_prob    = torch.softmax(outputs, dim=1)[:, 2]
        binary_preds = (mono_prob >= threshold).long()
        binary_true  = torch.tensor(
            [COLLAPSE_MAP[l.item()] for l in labels], device=device
        )

        all_preds.extend(binary_preds.cpu().numpy())
        all_labels.extend(binary_true.cpu().numpy())

    acc = sum(p == l for p, l in zip(all_preds, all_labels)) / total
    return total_loss / total, acc, all_preds, all_labels

# ─────────────────────────────────────────────
# THRESHOLD SEARCH
# ─────────────────────────────────────────────
def threshold_search(model, test_loader, device):
    all_probs, all_labels = [], []
    model.eval()
    with torch.no_grad():
        for imgs, labels in test_loader:
            imgs = imgs.to(device)
            all_probs.extend(torch.softmax(model(imgs), dim=1)[:, 2].cpu().numpy())
            all_labels.extend([COLLAPSE_MAP[l.item()] for l in labels])

    print(f"\n── Threshold Search ──")
    print(f"{'Threshold':>10} {'Mono Recall':>12} {'Mono Prec':>10} "
          f"{'NonMono Recall':>15} {'NonMono Prec':>13} {'Accuracy':>10}")
    print("-" * 75)

    for t in np.arange(0.20, 0.81, 0.05):
        preds        = (np.array(all_probs) >= t).astype(int)
        prec, rec, _, _ = precision_recall_fscore_support(
            all_labels, preds, labels=[0, 1], zero_division=0
        )
        acc = np.mean(preds == np.array(all_labels))
        print(f"{t:>10.2f} {rec[1]:>12.4f} {prec[1]:>10.4f} "
              f"{rec[0]:>15.4f} {prec[0]:>13.4f} {acc:>10.4f}")

    print(f"\n  Set classification_threshold in config manually.")

# ─────────────────────────────────────────────
# MULTICLASS DIAGNOSTIC
# ─────────────────────────────────────────────
def evaluate_multiclass(model, loader, device, output_dir):
    model.eval()
    all_preds, all_labels = [], []
    with torch.no_grad():
        for imgs, labels in loader:
            imgs = imgs.to(device)
            all_preds.extend(model(imgs).argmax(1).cpu().numpy())
            all_labels.extend(labels.cpu().numpy())

    cm = confusion_matrix(all_labels, all_preds, labels=[0, 1, 2])
    print("\n── 3-Class Confusion Matrix (raw counts) ──")
    print(cm)

    plt.figure(figsize=(8, 6))
    sns.heatmap(cm, annot=True, fmt="d", cmap="Blues",
                xticklabels=["Unusable", "RBCalone", "Monocyte"],
                yticklabels=["Unusable", "RBCalone", "Monocyte"])
    plt.title("3-Class Confusion Matrix (counts)")
    plt.ylabel("True Label")
    plt.xlabel("Predicted Label")
    plt.tight_layout()
    path = os.path.join(output_dir, "confusion_matrix_3class.png")
    plt.savefig(path, dpi=150)
    plt.close()
    print(f"  ✓ Saved {path}")

# ─────────────────────────────────────────────
# PLOTS
# ─────────────────────────────────────────────
def save_plots(train_losses, val_losses, train_accs, val_accs,
               true_labels, preds, output_dir):
    os.makedirs(output_dir, exist_ok=True)

    plt.figure(figsize=(10, 5))
    plt.plot(train_losses, label="Train Loss", marker="o", markersize=3)
    plt.plot(val_losses,   label="Val Loss",   marker="o", markersize=3)
    plt.title("Training & Validation Loss")
    plt.xlabel("Epoch")
    plt.ylabel("Loss")
    plt.legend()
    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, "loss_curve.png"), dpi=150)
    plt.close()

    plt.figure(figsize=(10, 5))
    plt.plot(train_accs, label="Train Accuracy (3-class)", marker="o", markersize=3)
    plt.plot(val_accs,   label="Val Accuracy (binary)",    marker="o", markersize=3)
    plt.title("Training & Validation Accuracy")
    plt.xlabel("Epoch")
    plt.ylabel("Accuracy")
    plt.ylim(0, 1)
    plt.legend()
    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, "accuracy_curve.png"), dpi=150)
    plt.close()

    print("  ✓ Saved loss_curve.png and accuracy_curve.png")

# ─────────────────────────────────────────────
# EXCEL LOGGING
# ─────────────────────────────────────────────
def log_experiment_excel(config, results,
                         log_path="C:/repos/Blood-Cell-Classification/experiment_log.xlsx",
                         notes=""):
    wb      = load_workbook(log_path)
    ws      = wb.active
    next_row = ws.max_row + 1
    run_num  = next_row - 2
    targets  = config["subclass_targets"]

    nonmono = sum(targets.get(k, 0) for k in ["Unusable", "RBCalone"])
    mono    = sum(targets.get(k, 0) for k in ["MCwRBC", "MCwoRBC", "Clustered"])

    row = [
        run_num, str(date.today()), notes,
        config.get("architecture", "resnet34"), "ImageNet", "No", 0.4, 12494,
        70, 10, 20, nonmono, mono,
        targets.get("MCwRBC", 0), targets.get("MCwoRBC", 0), targets.get("Clustered", 0),
        "Per-subclass inverse frequency", "AdamW",
        config["lr"], config["weight_decay"], config["batch_size"], config["num_epochs"],
        f"CosineAnnealingLR (T_max={config['num_epochs']})",
        f"{config['img_size']}x{config['img_size']}",
        "HFlip, VFlip, Rotation(15°), ColorJitter(b=0.2,c=0.2), ImageNet norm",
        results["test_acc"],
        results["non_mono_prec"], results["non_mono_recall"], results["non_mono_f1"],
        results["mono_prec"],     results["mono_recall"],     results["mono_f1"],
        results["macro_f1"],      results["weighted_f1"],
        results["tp"], results["tn"], results["fp"], results["fn"],
        results["val_acc"],
    ]

    for col_idx, value in enumerate(row, start=1):
        ws.cell(row=next_row, column=col_idx, value=value)

    wb.save(log_path)
    print(f"  ✓ Run {run_num} logged to {log_path}")

# ─────────────────────────────────────────────
# MAIN TRAIN FUNCTION
# ─────────────────────────────────────────────
def train(config=None, notes=""):
    """
    Train Stage 1 classifier.

    Args:
        config: dict — override any keys from DEFAULT_CONFIG.
                Pass None to use DEFAULT_CONFIG unchanged.
        notes:  string — description logged to MLflow and Excel.

    Example:
        from classifier_stage1 import train, DEFAULT_CONFIG
        import copy

        config = copy.deepcopy(DEFAULT_CONFIG)
        config["architecture"]    = "resnet50"
        config["checkpoint_path"] = "checkpoints_stage1/resnet50.pth"
        config["subclass_targets"]["MCwRBC"] = 900

        train(config, notes="resnet50 higher MCwRBC")
    """
    # Merge provided config over defaults
    cfg = copy.deepcopy(DEFAULT_CONFIG)
    if config is not None:
        for k, v in config.items():
            if k == "subclass_targets" and isinstance(v, dict):
                cfg["subclass_targets"].update(v)  # merge targets, not replace
            else:
                cfg[k] = v

    device = cfg["device"]
    arch   = cfg.get("architecture", "resnet34")
    print(f"\nTraining on: {device}  |  Architecture: {arch}")

    os.makedirs(os.path.dirname(cfg["checkpoint_path"]), exist_ok=True)
    plot_dir = os.path.dirname(cfg["checkpoint_path"])

    # MLflow
    setup_mlflow("Stage1_MCvsNonMC")
    start_run(run_name=f"{arch}_{notes}" if notes else arch, notes=notes)
    log_params(cfg)

    train_loader, val_loader, test_loader, fine_counts = make_dataloaders(cfg)
    model     = build_model(arch, num_classes=3, freeze_backbone=False).to(device)
    criterion = nn.CrossEntropyLoss()
    optimizer = torch.optim.AdamW(model.parameters(), lr=cfg["lr"], weight_decay=cfg["weight_decay"])
    scheduler = torch.optim.lr_scheduler.CosineAnnealingLR(optimizer, T_max=cfg["num_epochs"])

    train_losses, val_losses = [], []
    train_accs,   val_accs   = [], []
    best_val_acc              = 0.0
    epochs_no_improve         = 0

    print("\n── Training ──")
    for epoch in range(1, cfg["num_epochs"] + 1):
        train_loss, train_acc = _train_one_epoch_clean(model, train_loader, optimizer, criterion, device)
        val_loss,   val_acc, _, _ = evaluate(model, val_loader, criterion, device, cfg["classification_threshold"])
        scheduler.step()

        train_losses.append(train_loss)
        val_losses.append(val_loss)
        train_accs.append(train_acc)
        val_accs.append(val_acc)

        print(f"Epoch {epoch:02d}/{cfg['num_epochs']} | "
              f"Train Loss: {train_loss:.4f}  Acc: {train_acc:.4f} | "
              f"Val Loss: {val_loss:.4f}  Acc: {val_acc:.4f}")

        if val_acc > best_val_acc:
            best_val_acc      = val_acc
            epochs_no_improve = 0
            torch.save(model.state_dict(), cfg["checkpoint_path"])
            print(f"  ✓ New best model saved (val_acc={val_acc:.4f})")
        else:
            epochs_no_improve += 1
            print(f"  No improvement ({epochs_no_improve}/{cfg['early_stopping_patience']})")
            if epochs_no_improve >= cfg["early_stopping_patience"]:
                print(f"\n  Early stopping at epoch {epoch}")
                break

        log_epoch(epoch, train_loss, train_acc, val_loss, val_acc)

    # ── Load best checkpoint ──
    model.load_state_dict(torch.load(cfg["checkpoint_path"]))

    # ── Diagnostics ──
    evaluate_multiclass(model, test_loader, device, plot_dir)
    threshold_search(model, test_loader, device)

    # ── Val evaluation ──
    print("\n── Validation Set Evaluation ──")
    _, val_acc_final, val_preds, val_true = evaluate(
        model, val_loader, criterion, device, cfg["classification_threshold"]
    )
    print(f"Val Accuracy: {val_acc_final:.4f}\n")
    print(classification_report(val_true, val_preds, target_names=["Non-Monocyte", "Monocyte"]))
    print(confusion_matrix(val_true, val_preds))

    # ── Test evaluation ──
    print("\n── Test Set Evaluation ──")
    _, test_acc, preds, true_labels = evaluate(
        model, test_loader, criterion, device, cfg["classification_threshold"]
    )
    print(f"Test Accuracy: {test_acc:.4f}\n")
    print(classification_report(true_labels, preds, target_names=["Non-Monocyte", "Monocyte"]))
    print(confusion_matrix(true_labels, preds))

    # ── Save plots ──
    save_plots(train_losses, val_losses, train_accs, val_accs, true_labels, preds, plot_dir)

    # ── Compile results ──
    prec, rec, f1, _ = precision_recall_fscore_support(true_labels, preds, labels=[0, 1])
    _, _, f1w, _     = precision_recall_fscore_support(true_labels, preds, average="weighted")
    _, _, f1m, _     = precision_recall_fscore_support(true_labels, preds, average="macro")
    cm               = confusion_matrix(true_labels, preds)

    results = {
        "test_acc":        round(test_acc,   4),
        "non_mono_prec":   round(prec[0],    4),
        "non_mono_recall": round(rec[0],     4),
        "non_mono_f1":     round(f1[0],      4),
        "mono_prec":       round(prec[1],    4),
        "mono_recall":     round(rec[1],     4),
        "mono_f1":         round(f1[1],      4),
        "macro_f1":        round(f1m,        4),
        "weighted_f1":     round(f1w,        4),
        "tp": int(cm[1,1]), "tn": int(cm[0,0]),
        "fp": int(cm[0,1]), "fn": int(cm[1,0]),
        "val_acc": round(val_acc_final, 4),
    }

    # ── Log everything ──
    log_experiment_excel(cfg, results, notes=notes)
    log_results(results)
    log_confusion_matrix(true_labels, preds, ["Non-Monocyte", "Monocyte"],
                         os.path.join(plot_dir, "confusion_matrix.png"))
    log_artifacts([
        os.path.join(plot_dir, "loss_curve.png"),
        os.path.join(plot_dir, "accuracy_curve.png"),
        os.path.join(plot_dir, "confusion_matrix_3class.png"),
        cfg["checkpoint_path"],
    ])
    end_run()

    return results


# ─────────────────────────────────────────────
# ENTRY POINT — run directly with default config
# ─────────────────────────────────────────────
if __name__ == "__main__":
    train(notes="default config run")