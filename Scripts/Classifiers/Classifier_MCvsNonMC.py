import os
import torch
import torch.nn as nn
from torch.utils.data import Dataset, DataLoader, WeightedRandomSampler
from torchvision import transforms, models
from PIL import Image
from sklearn.model_selection import train_test_split
from sklearn.metrics import classification_report, confusion_matrix
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import load_workbook
from datetime import date
from sklearn.metrics import precision_recall_fscore_support

# ─────────────────────────────────────────────
# CONFIG
#
# 3-class training collapsed to binary at inference:
#   Class 0 — Unusable + Lymphocyte (labelled interchangeably)
#   Class 1 — RBC alone
#   Class 2 — Monocyte (MCwRBC + MCwoRBC + Clustered)
#
# At inference: class 0 + class 1 → NonMonocyte, class 2 → Monocyte
# ─────────────────────────────────────────────
_mono_total = 2400  # 737 × 3

CONFIG = {
    "data_dir": "D:/MMA_LabelledData/Sliced",
    "img_size": 256,
    "batch_size": 32,
    "num_epochs": 20,
    "early_stopping_patience": 5,
    "lr": 5e-5,
    "weight_decay": 1e-4,
    "num_workers": 4,
    "checkpoint_path": "checkpoints_stage1/resnet34_binary.pth",
    "device": "cuda" if torch.cuda.is_available() else "cpu",
    "classification_threshold": 0.45,  # applied to Monocyte class (class 2) probability

    # Per-subclass sampling targets
    # NonMono total: 600 + 300 + 300 = 1200 per epoch
    # Monocyte total: 737 × 3 = 2211 per epoch
    # Duplication rates:
    #   Unusable   (~1241) → 600:  0.48x subsampled
    #   Lymphocyte  (~153) → 300:  ~2x oversample (mild)
    #   RBCalone    (~273) → 300:  ~1.1x essentially natural
    #   MCwRBC      (~515) → 737:  ~1.4x mild oversample
    #   MCwoRBC    (~2764) → 737:  0.27x subsampled
    #   Clustered  (~3748) → 737:  0.20x subsampled
    "subclass_targets": {
        "Unusable":   800,
        "Lymphocyte": 200,
        "RBCalone":   400,
        "MCwRBC":     _mono_total // 3,
        "MCwoRBC":    _mono_total // 3,
        "Clustered":  _mono_total - 2 * (_mono_total // 3),
    }
}

# ─────────────────────────────────────────────
# LABEL MAPS
#
# FINE_MAP:     folder name  → fine label
# LABEL_MAP:    fine label   → training class (0, 1, 2)
# COLLAPSE_MAP: training class → binary (0=NonMono, 1=Monocyte)
# ─────────────────────────────────────────────
FINE_MAP = {
    "Monocyte_with_RBC":    "MCwRBC",
    "Monocyte_without_RBC": "MCwoRBC",
    "Clustered_cell":       "Clustered",
    "Unusable":             "Unusable",
    "Lymphocyte":           "Lymphocyte",
    "RBC alone":            "RBCalone",
}

LABEL_MAP = {
    "Unusable":   0,   # NonMono subclass 1
    "Lymphocyte": 0,   # lumped with Unusable — labelled interchangeably
    "RBCalone":   1,   # NonMono subclass 2
    "MCwRBC":     2,   # Monocyte
    "MCwoRBC":    2,   # Monocyte
    "Clustered":  2,   # Monocyte
}

COLLAPSE_MAP = {
    0: 0,  # Unusable + Lymphocyte → NonMonocyte
    1: 0,  # RBCalone              → NonMonocyte
    2: 1,  # Monocyte              → Monocyte
}

# ─────────────────────────────────────────────
# DATASET
# ─────────────────────────────────────────────
class BloodCellDataset(Dataset):
    def __init__(self, samples, transform=None):
        """
        samples: list of (image_path, training_label) tuples
        training_label is 0/1/2 (3-class)
        """
        self.samples = samples
        self.transform = transform

    def __len__(self):
        return len(self.samples)

    def __getitem__(self, idx):
        path, label = self.samples[idx]
        img = Image.open(path).convert("RGB")
        if self.transform:
            img = self.transform(img)
        return img, label

# ─────────────────────────────────────────────
# TRANSFORMS — light augmentation (Run 13 baseline)
# ─────────────────────────────────────────────
train_transforms = transforms.Compose([
    transforms.Resize((256, 256)),
    transforms.RandomHorizontalFlip(),
    transforms.RandomVerticalFlip(),
    transforms.RandomRotation(15),
    transforms.ColorJitter(brightness=0.2, contrast=0.2),
    transforms.ToTensor(),
    transforms.Normalize(mean=[0.485, 0.456, 0.406],
                         std=[0.229, 0.224, 0.225]),
])

val_transforms = transforms.Compose([
    transforms.Resize((256, 256)),
    transforms.ToTensor(),
    transforms.Normalize(mean=[0.485, 0.456, 0.406],
                         std=[0.229, 0.224, 0.225]),
])

# ─────────────────────────────────────────────
# DATA LOADING
# ─────────────────────────────────────────────
def load_samples(data_dir, fine_map, label_map):
    """
    Walk subfolders and collect (path, fine_label, training_label) tuples.
    fine_label used for per-subclass sampling.
    training_label (0/1/2) used for model training.
    """
    samples = []
    for subfolder, fine_label in fine_map.items():
        folder_path = os.path.join(data_dir, subfolder)
        if not os.path.isdir(folder_path):
            print(f"Warning: folder not found — {folder_path}")
            continue
        training_label = label_map[fine_label]
        for fname in os.listdir(folder_path):
            if fname.lower().endswith(('.png', '.jpg', '.jpeg')):
                samples.append((
                    os.path.join(folder_path, fname),
                    fine_label,
                    training_label
                ))
    return samples


def make_weighted_sampler(train_samples, subclass_targets):
    """
    Per-subclass sampling — each fine class drawn at its own target rate.
    Duplication capped at ~2x for smallest classes.
    """
    fine_counts = {}
    for _, fine_label, _ in train_samples:
        fine_counts[fine_label] = fine_counts.get(fine_label, 0) + 1

    print("\nNatural class counts in training set:")
    for label, count in sorted(fine_counts.items()):
        target = subclass_targets.get(label, count)
        dup = target / count
        binary = "Monocyte" if LABEL_MAP.get(label, 0) == 2 else "Non-Monocyte"
        print(f"  {label} ({binary}): {count} → {target} ({dup:.2f}x)")

    sample_weights = np.zeros(len(train_samples), dtype=np.float32)
    for idx, (_, fine_label, _) in enumerate(train_samples):
        natural = fine_counts[fine_label]
        target  = subclass_targets.get(fine_label, natural)
        sample_weights[idx] = target / natural

    total_samples = sum(subclass_targets.values())
    sampler = WeightedRandomSampler(
        weights=torch.tensor(sample_weights),
        num_samples=total_samples,
        replacement=True
    )

    nonmono_total = (subclass_targets.get("Unusable", 0) +
                     subclass_targets.get("Lymphocyte", 0) +
                     subclass_targets.get("RBCalone", 0))
    mono_total    = (subclass_targets.get("MCwRBC", 0) +
                     subclass_targets.get("MCwoRBC", 0) +
                     subclass_targets.get("Clustered", 0))

    print(f"\nEffective samples per epoch:")
    print(f"  NonMonocyte total: {nonmono_total}")
    print(f"  Monocyte total:    {mono_total}")
    print(f"  Grand total:       {total_samples}")

    return sampler, fine_counts


def make_dataloaders(config):
    all_samples = load_samples(config["data_dir"], FINE_MAP, LABEL_MAP)

    # Stratify on binary label to preserve Mono/NonMono ratio in splits
    binary_labels = [COLLAPSE_MAP[s[2]] for s in all_samples]

    # 70% train, 30% temp
    train, temp = train_test_split(
        all_samples, test_size=0.30,
        stratify=binary_labels, random_state=42
    )

    # Split temp → 20% test, 10% val
    temp_binary = [COLLAPSE_MAP[s[2]] for s in temp]
    test, val = train_test_split(
        temp, test_size=0.333,
        stratify=temp_binary, random_state=42
    )

    print(f"\nDataset split:")
    print(f"  Train: {len(train)} ({len(train)/len(all_samples)*100:.1f}%)")
    print(f"  Test:  {len(test)}  ({len(test)/len(all_samples)*100:.1f}%)")
    print(f"  Val:   {len(val)}   ({len(val)/len(all_samples)*100:.1f}%)")

    sampler, fine_counts = make_weighted_sampler(
        train, config["subclass_targets"]
    )

    # Use 3-class training labels for dataset
    train_ds = BloodCellDataset(
        [(path, label) for path, _, label in train],
        transform=train_transforms
    )
    val_ds = BloodCellDataset(
        [(path, label) for path, _, label in val],
        transform=val_transforms
    )
    test_ds = BloodCellDataset(
        [(path, label) for path, _, label in test],
        transform=val_transforms
    )

    train_loader = DataLoader(
        train_ds, batch_size=config["batch_size"],
        sampler=sampler, num_workers=config["num_workers"], pin_memory=True
    )
    val_loader = DataLoader(
        val_ds, batch_size=config["batch_size"],
        shuffle=False, num_workers=config["num_workers"], pin_memory=True
    )
    test_loader = DataLoader(
        test_ds, batch_size=config["batch_size"],
        shuffle=False, num_workers=config["num_workers"], pin_memory=True
    )

    return train_loader, val_loader, test_loader, fine_counts

# ─────────────────────────────────────────────
# MODEL — 3 output classes
# ─────────────────────────────────────────────
def build_model(num_classes=3, freeze_backbone=False):
    model = models.resnet34(weights=models.ResNet34_Weights.DEFAULT)

    if freeze_backbone:
        for param in model.parameters():
            param.requires_grad = False

    model.fc = nn.Sequential(
        nn.Dropout(p=0.4),
        nn.Linear(model.fc.in_features, num_classes)
    )
    return model

# ─────────────────────────────────────────────
# TRAINING LOOP
# ─────────────────────────────────────────────
def train_one_epoch(model, loader, optimizer, criterion, device):
    model.train()
    total_loss, correct, total = 0, 0, 0
    for imgs, labels in loader:
        imgs, labels = imgs.to(device), labels.to(device)
        optimizer.zero_grad()
        outputs = model(imgs)
        loss = criterion(outputs, labels)
        loss.backward()
        optimizer.step()
        total_loss += loss.item() * imgs.size(0)
        # 3-class accuracy during training
        correct += (outputs.argmax(1) == labels).sum().item()
        total += imgs.size(0)
    return total_loss / total, correct / total

# ─────────────────────────────────────────────
# EVALUATE
# Runs 3-class model, collapses to binary for reporting.
# Threshold applied to Monocyte (class 2) probability.
# ─────────────────────────────────────────────
@torch.no_grad()
def evaluate(model, loader, criterion, device, threshold=0.5):
    model.eval()
    total_loss, total = 0, 0
    all_preds, all_labels = [], []

    for imgs, labels in loader:
        imgs, labels = imgs.to(device), labels.to(device)
        outputs = model(imgs)
        loss = criterion(outputs, labels)
        total_loss += loss.item() * imgs.size(0)
        total += imgs.size(0)

        # Monocyte probability = softmax score of class 2
        probs = torch.softmax(outputs, dim=1)
        mono_prob = probs[:, 2]

        # Binary prediction via threshold
        binary_preds = (mono_prob >= threshold).long()

        # Collapse true 3-class labels to binary
        binary_labels = torch.tensor(
            [COLLAPSE_MAP[l.item()] for l in labels],
            device=device
        )

        all_preds.extend(binary_preds.cpu().numpy())
        all_labels.extend(binary_labels.cpu().numpy())

    correct = sum(p == l for p, l in zip(all_preds, all_labels))
    acc = correct / total

    return total_loss / total, acc, all_preds, all_labels

# Determine the confusion matrix for all separate classes
def evaluate_multiclass(model, loader, device):
    """
    Evaluate raw 3-class predictions without collapsing to binary.
    Useful for diagnosing which NonMono subclass is being confused with Monocyte.
    """
    model.eval()
    all_preds, all_labels = [], []
    with torch.no_grad():
        for imgs, labels in loader:
            imgs = imgs.to(device)
            outputs = model(imgs)
            preds = outputs.argmax(1)
            all_preds.extend(preds.cpu().numpy())
            all_labels.extend(labels.cpu().numpy())

    print("\n── 3-Class Confusion Matrix ──")
    print("(Rows=True, Cols=Predicted)")
    cm = confusion_matrix(all_labels, all_preds, labels=[0, 1, 2])
    cm_percent = cm.astype(float) / cm.sum(axis=1, keepdims=True) * 100
    plt.figure(figsize=(8, 6))
    sns.heatmap(cm_percent, annot=True, fmt=".1f", cmap="Blues",
                xticklabels=["Unusable/Lymphocyte", "RBCalone", "Monocyte"],
                yticklabels=["Unusable/Lymphocyte", "RBCalone", "Monocyte"])
    plt.title("3-Class Confusion Matrix (%)")
    plt.ylabel("True Label")
    plt.xlabel("Predicted Label")
    for text in plt.gca().texts:
        text.set_text(text.get_text() + "%")
    plt.tight_layout()
    plt.savefig("checkpoints_stage1/confusion_matrix_3class.png", dpi=150)
    plt.close()
    print(cm)
    print("  ✓ Saved confusion_matrix_3class.png")



# ─────────────────────────────────────────────
# THRESHOLD SEARCH
# Prints full tradeoff table — set threshold manually in CONFIG
# ─────────────────────────────────────────────
def threshold_search(model, test_loader, criterion, device):
    all_mono_probs, all_binary_labels = [], []
    model.eval()
    with torch.no_grad():
        for imgs, labels in test_loader:
            imgs = imgs.to(device)
            probs = torch.softmax(model(imgs), dim=1)
            all_mono_probs.extend(probs[:, 2].cpu().numpy())
            all_binary_labels.extend(
                [COLLAPSE_MAP[l.item()] for l in labels]
            )

    print(f"\n── Threshold Search ──")
    print(f"{'Threshold':>10} {'Mono Recall':>12} {'Mono Prec':>10} "
          f"{'NonMono Recall':>15} {'NonMono Prec':>13} {'Accuracy':>10}")
    print("-" * 75)

    for threshold in np.arange(0.20, 0.81, 0.05):
        preds = (np.array(all_mono_probs) >= threshold).astype(int)
        prec, rec, _, _ = precision_recall_fscore_support(
            all_binary_labels, preds, labels=[0, 1], zero_division=0
        )
        acc = np.mean(np.array(preds) == np.array(all_binary_labels))
        print(f"{threshold:>10.2f} {rec[1]:>12.4f} {prec[1]:>10.4f} "
              f"{rec[0]:>15.4f} {prec[0]:>13.4f} {acc:>10.4f}")

    print(f"\n  Review table above and set classification_threshold in CONFIG manually.")

# ─────────────────────────────────────────────
# PLOTTING
# ─────────────────────────────────────────────
def save_plots(train_losses, val_losses, train_accs, val_accs,
               true_labels, preds, output_dir="checkpoints_stage1"):
    os.makedirs(output_dir, exist_ok=True)

    # ── Loss curve ──
    plt.figure(figsize=(10, 5))
    plt.plot(train_losses, label="Train Loss", marker="o", markersize=3)
    plt.plot(val_losses,   label="Val Loss",   marker="o", markersize=3)
    plt.title("Training & Validation Loss")
    plt.xlabel("Epoch")
    plt.ylabel("Loss")
    plt.legend()
    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, "loss_curve.png"), dpi=150)
    plt.close()
    print("  ✓ Saved loss_curve.png")

    # ── Accuracy curve ──
    plt.figure(figsize=(10, 5))
    plt.plot(train_accs, label="Train Accuracy (3-class)", marker="o", markersize=3)
    plt.plot(val_accs,   label="Val Accuracy (binary)",    marker="o", markersize=3)
    plt.title("Training & Validation Accuracy")
    plt.xlabel("Epoch")
    plt.ylabel("Accuracy")
    plt.ylim(0, 1)
    plt.legend()
    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, "accuracy_curve.png"), dpi=150)
    plt.close()
    print("  ✓ Saved accuracy_curve.png")

    # ── Confusion matrix heatmap (%) ──
    cm = confusion_matrix(true_labels, preds)
    cm_percent = cm.astype(float) / cm.sum(axis=1, keepdims=True) * 100
    plt.figure(figsize=(7, 6))
    sns.heatmap(cm_percent, annot=True, fmt=".1f", cmap="Blues",
                xticklabels=["Non-Monocyte", "Monocyte"],
                yticklabels=["Non-Monocyte", "Monocyte"])
    plt.title("Confusion Matrix (%)")
    plt.ylabel("True Label")
    plt.xlabel("Predicted Label")
    for text in plt.gca().texts:
        text.set_text(text.get_text() + "%")
    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, "confusion_matrix.png"), dpi=150)
    plt.close()
    print("  ✓ Saved confusion_matrix.png")

# ─────────────────────────────────────────────
# EXPERIMENT LOGGING
# ─────────────────────────────────────────────
def log_experiment(config, results, log_path="experiment_log.xlsx", notes=""):
    wb = load_workbook(log_path)
    ws = wb.active

    next_row = ws.max_row + 1
    run_num  = next_row - 2

    aug_str  = "HFlip, VFlip, Rotation(15°), ColorJitter(b=0.2,c=0.2), ImageNet norm"
    targets  = config["subclass_targets"]

    nonmono_total = (targets.get("Unusable", 0) +
                     targets.get("Lymphocyte", 0) +
                     targets.get("RBCalone", 0))
    mono_total    = (targets.get("MCwRBC", 0) +
                     targets.get("MCwoRBC", 0) +
                     targets.get("Clustered", 0))

    row = [
        run_num,
        str(date.today()),
        notes,
        "ResNet34", "ImageNet",
        "No",    # freeze_backbone
        0.4,     # dropout
        12494,
        70, 10, 20,
        nonmono_total,
        mono_total,
        targets.get("MCwRBC", 0),
        targets.get("MCwoRBC", 0),
        targets.get("Clustered", 0),
        "Per-subclass inverse frequency (3-class training)",
        "AdamW",
        config["lr"],
        config["weight_decay"],
        config["batch_size"],
        config["num_epochs"],
        f"CosineAnnealingLR (T_max={config['num_epochs']})",
        f"{config['img_size']}x{config['img_size']}",
        aug_str,
        results["test_acc"],
        results["non_mono_prec"],
        results["non_mono_recall"],
        results["non_mono_f1"],
        results["mono_prec"],
        results["mono_recall"],
        results["mono_f1"],
        results["macro_f1"],
        results["weighted_f1"],
        results["tp"],
        results["tn"],
        results["fp"],
        results["fn"],
        results["val_acc"],
    ]

    for col_idx, value in enumerate(row, start=1):
        ws.cell(row=next_row, column=col_idx, value=value)

    wb.save(log_path)
    print(f"  ✓ Run {run_num} logged to {log_path}")

# ─────────────────────────────────────────────
# TRAIN
# ─────────────────────────────────────────────
def train(config):
    device = config["device"]
    print(f"Training on: {device}")
    os.makedirs(os.path.dirname(config["checkpoint_path"]), exist_ok=True)

    train_loader, val_loader, test_loader, fine_counts = make_dataloaders(config)

    # 3-class model — no backbone freeze (consistent with Run 13 baseline)
    model = build_model(num_classes=3, freeze_backbone=False).to(device)

    # No class weights — sampler already balances subclasses
    criterion = nn.CrossEntropyLoss()

    optimizer = torch.optim.AdamW(
        model.parameters(), lr=config["lr"], weight_decay=config["weight_decay"]
    )
    scheduler = torch.optim.lr_scheduler.CosineAnnealingLR(
        optimizer, T_max=config["num_epochs"]
    )

    # ── Track metrics per epoch ──
    train_losses, val_losses = [], []
    train_accs,   val_accs   = [], []
    best_val_acc  = 0.0
    epochs_without_improvement = 0

    print("\n── Training ──")
    for epoch in range(1, config["num_epochs"] + 1):

        train_loss, train_acc = train_one_epoch(
            model, train_loader, optimizer, criterion, device
        )
        val_loss, val_acc, _, _ = evaluate(
            model, val_loader, criterion, device,
            threshold=config["classification_threshold"]
        )
        scheduler.step()

        train_losses.append(train_loss)
        val_losses.append(val_loss)
        train_accs.append(train_acc)
        val_accs.append(val_acc)

        print(f"Epoch {epoch:02d}/{config['num_epochs']} | "
              f"Train Loss: {train_loss:.4f}  Acc: {train_acc:.4f} | "
              f"Val Loss: {val_loss:.4f}  Acc: {val_acc:.4f}")

        # Save best model by binary val accuracy
        if val_acc > best_val_acc:
            best_val_acc = val_acc
            torch.save(model.state_dict(), config["checkpoint_path"])
            epochs_without_improvement = 0
            print(f"  ✓ New best model saved (val_acc={val_acc:.4f})")
        else:
            epochs_without_improvement += 1
            print(f"  No improvement ({epochs_without_improvement}/{config['early_stopping_patience']})")
            if epochs_without_improvement >= config["early_stopping_patience"]:
                print(f"\n  Early stopping triggered at epoch {epoch}")
                break

    # ── Load best checkpoint ──
    print("\n── Loading Best Checkpoint ──")
    model.load_state_dict(torch.load(config["checkpoint_path"]))    

    # ── 3-class diagnostic ──
    evaluate_multiclass(model, test_loader, device)

    # ── Threshold search — review and set manually in CONFIG ──
    threshold_search(model, test_loader, criterion, device)

    # ── Validation set evaluation ──
    print("\n── Validation Set Evaluation ──")
    _, val_acc_final, val_preds, val_true = evaluate(
        model, val_loader, criterion, device,
        threshold=config["classification_threshold"]
    )
    print(f"Val Accuracy: {val_acc_final:.4f}\n")
    print(classification_report(
        val_true, val_preds,
        target_names=["Non-Monocyte", "Monocyte"]
    ))
    print("Confusion Matrix:")
    print(confusion_matrix(val_true, val_preds))

    # ── Test set evaluation ──
    print("\n── Test Set Evaluation ──")
    _, test_acc, preds, true_labels = evaluate(
        model, test_loader, criterion, device,
        threshold=config["classification_threshold"]
    )
    print(f"Test Accuracy: {test_acc:.4f}\n")
    print(classification_report(
        true_labels, preds,
        target_names=["Non-Monocyte", "Monocyte"]
    ))
    print("Confusion Matrix:")
    print(confusion_matrix(true_labels, preds))

    # ── Save plots ──
    print("\n── Saving Plots ──")
    plot_dir = os.path.dirname(config["checkpoint_path"])
    save_plots(train_losses, val_losses, train_accs, val_accs,
               true_labels, preds, output_dir=plot_dir)

    # ── Log experiment ──
    prec, rec, f1, _ = precision_recall_fscore_support(
        true_labels, preds, labels=[0, 1]
    )
    _, _, f1_weighted, _ = precision_recall_fscore_support(
        true_labels, preds, average="weighted"
    )
    _, _, f1_macro, _ = precision_recall_fscore_support(
        true_labels, preds, average="macro"
    )
    cm = confusion_matrix(true_labels, preds)

    log_experiment(config, {
        "test_acc":        round(test_acc, 4),
        "non_mono_prec":   round(prec[0], 4),
        "non_mono_recall": round(rec[0],  4),
        "non_mono_f1":     round(f1[0],   4),
        "mono_prec":       round(prec[1], 4),
        "mono_recall":     round(rec[1],  4),
        "mono_f1":         round(f1[1],   4),
        "macro_f1":        round(f1_macro,    4),
        "weighted_f1":     round(f1_weighted, 4),
        "tp": int(cm[1,1]), "tn": int(cm[0,0]),
        "fp": int(cm[0,1]), "fn": int(cm[1,0]),
        "val_acc": round(val_acc_final, 4),
    }, log_path="C:/repos/Blood-Cell-Classification/experiment_log.xlsx",
       notes="3-class training: Unusable+Lymphocyte / RBCalone / Monocyte — collapsed to binary")

# ─────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────
if __name__ == "__main__":
    train(CONFIG)